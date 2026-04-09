from __future__ import annotations

from collections.abc import Iterable, Sequence
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


class ExportService:
    def __init__(self) -> None:
        self._aggregate_green_fill = PatternFill(fill_type="solid", fgColor="DAF2D0")
        self._aggregate_orange_fill = PatternFill(fill_type="solid", fgColor="FBE2D5")

    def export_dataframe(
        self,
        dataframe: pd.DataFrame,
        target_path: Path,
        formatter: str | None = None,
    ) -> int:
        rows = dataframe.itertuples(index=False, name=None)
        return self.export_rows(list(dataframe.columns), rows, target_path, formatter=formatter)

    def export_rows(
        self,
        columns: Sequence[str],
        rows: Iterable[Sequence[object]],
        target_path: Path,
        formatter: str | None = None,
    ) -> int:
        workbook = Workbook(write_only=True)
        sheet = workbook.create_sheet(title="Sheet1")
        self._apply_sheet_layout(sheet, formatter)

        widths = [len(str(column)) + 2 for column in columns]
        sheet.append(list(columns))

        row_count = 0
        for row in rows:
            values = list(row)
            row_count += 1
            for index, value in enumerate(values):
                widths[index] = max(widths[index], min(len(str(value or "")) + 2, 40))
            sheet.append(self._render_row(sheet, values, formatter))

        self._apply_widths(sheet, widths)
        if row_count > 0:
            sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{row_count + 1}"
        workbook.save(target_path)
        return row_count

    def _apply_sheet_layout(self, sheet, formatter: str | None) -> None:  # type: ignore[no-untyped-def]
        sheet.freeze_panes = "A2"
        if formatter == "aggregate":
            for column_letter in ["E", "F", "G", "H", "I"]:
                sheet.column_dimensions[column_letter].hidden = True

    def _render_row(self, sheet, values: list[object], formatter: str | None) -> list[object]:  # type: ignore[no-untyped-def]
        if formatter != "aggregate":
            return values

        rendered: list[object] = []
        for index, value in enumerate(values):
            if index not in {14, 15, 18}:
                rendered.append(value)
                continue

            cell = WriteOnlyCell(sheet, value=value)
            if index == 14:
                cell.fill = self._aggregate_green_fill
            elif index == 15:
                cell.fill = self._aggregate_orange_fill
            elif index == 18:
                cell.number_format = "0.00%"
            rendered.append(cell)
        return rendered

    def _apply_widths(self, sheet, widths: Sequence[int]) -> None:  # type: ignore[no-untyped-def]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = min(width, 40)
